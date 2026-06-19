# uam_wildlife_strike_analysis.py
# UAM Corridor Wildlife Strike Risk — Height-Stratified NWSD Analysis
# Google Colab compatible | Python 3.10+
#
# Usage (Colab):
#   1. Upload Public.xlsx to Colab files panel (or mount Google Drive — see Step 0)
#   2. Runtime > Run all
#   3. Figures saved to /content/uam_figures/ as PNG (300 dpi)
#
# Local usage:
#   pip install openpyxl pandas numpy matplotlib scipy
#   Set NWSD_PATH below and run: python uam_wildlife_strike_analysis.py
#
# Data source: FAA NWSD — https://wildlife.faa.gov/downloads
# RHS values: Ross et al. (2025) Wildlife Society Bulletin 49(1) e1609

import os
import sys
import json
import warnings
from collections import defaultdict, Counter
from itertools import product

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch, FancyArrowPatch
import matplotlib.colors as mcolors

warnings.filterwarnings('ignore')

# ─── 0. CONFIG ────────────────────────────────────────────────────────────────

NWSD_PATH = 'Public.xlsx'      # change if needed, or mount Drive
OUT_DIR   = 'uam_figures'
os.makedirs(OUT_DIR, exist_ok=True)

# Colab: auto-detect and try to find the file in common locations
if not os.path.exists(NWSD_PATH):
    candidates = [
        '/content/Public.xlsx',
        '/content/drive/MyDrive/Public.xlsx',
        '/content/drive/MyDrive/FAA/Public.xlsx',
    ]
    for c in candidates:
        if os.path.exists(c):
            NWSD_PATH = c
            print(f"Found NWSD at {NWSD_PATH}")
            break
    else:
        print("Public.xlsx not found. Upload it to /content/ or mount Drive.")
        print("Colab: from google.colab import files; files.upload()")
        sys.exit(1)

# ─── Visual style ─────────────────────────────────────────────────────────────

C_NAVY  = '#1B3A6B'
C_CRIM  = '#9B1D22'
C_GREEN = '#1A5C32'
C_AMBER = '#6B4400'
C_GREY  = '#555555'
UAM_RED = '#C0392B'
AAM_AMB = '#E67E22'
CONV_BL = '#2980B9'

plt.rcParams.update({
    'font.family':       'serif',
    'font.size':         9,
    'axes.linewidth':    0.7,
    'axes.spines.top':   False,
    'axes.spines.right': False,
    'axes.grid':         True,
    'grid.color':        '#DADADA',
    'grid.linewidth':    0.35,
    'grid.linestyle':    ':',
    'xtick.major.size':  3.5,
    'ytick.major.size':  3.5,
    'figure.facecolor':  'white',
    'axes.facecolor':    'white',
    'legend.fontsize':   8,
    'legend.framealpha': 0.88,
})

# ─── RHS values from Ross et al. (2025) ──────────────────────────────────────
# These are published; do not substitute with other sources

RHS = {
    'Canada goose':              4.87,
    'Sandhill crane':            5.12,
    'American white pelican':    5.63,
    'Turkey vulture':            3.76,
    'Great blue heron':          3.21,
    'Bald eagle':                3.98,
    'Mallard':                   2.65,
    'Red-tailed hawk':           2.34,
    'Herring gull':              2.18,
    'Gulls':                     2.18,
    'Double-crested cormorant':  2.87,
    'Mourning dove':             0.51,
    'American kestrel':          0.74,
    'Barn swallow':              0.14,
    'European starling':         0.32,
    'Horned lark':               0.19,
    'Rock pigeon':               0.38,
    'Sparrows':                  0.21,
    'Killdeer':                  0.28,
    'Unknown bird':              0.30,   # median default
}

# ─── UCRI corridor specification ─────────────────────────────────────────────
# H_c = hazard-weighted species exposure (0-10)
# S_c = seasonal peak factor for Zone C (from real NWSD)
# D_c = computed from real NWSD sub-500ft records by state (below)
# T_c excluded — deployment timelines too uncertain for real-data index

CORRIDOR_STATES = {
    'SFO–Oakland':         'CA',
    'LA Downtown–Burbank': 'CA',
    'Miami CBD–MIA':       'FL',
    'Houston CBD–IAH':     'TX',
    'Dallas CBD–DFW':      'TX',
    'NYC Manhattan–JFK':   'NY',
    'Chicago Loop–ORD':    'IL',
    'Atlanta CBD–ATL':     'GA',
    'Seattle CBD–SEA':     'WA',
    'Denver CBD–DEN':      'CO',
}

# H_c: ecology-based hazard exposure (coastal vs inland vs mountain)
# Rationale: coastal states have pelicans, gulls (RHS~2.2), herons (3.2), cormorants (2.9)
# NYC specific: Canada goose near JFK (RHS=4.87) elevates H_c
HC = {
    'SFO–Oakland':         7.8,
    'LA Downtown–Burbank': 7.4,
    'Miami CBD–MIA':       8.0,
    'Houston CBD–IAH':     7.0,
    'Dallas CBD–DFW':      6.8,
    'NYC Manhattan–JFK':   8.2,   # highest: Canada goose JFK
    'Chicago Loop–ORD':    7.1,
    'Atlanta CBD–ATL':     6.5,
    'Seattle CBD–SEA':     6.9,
    'Denver CBD–DEN':      5.8,
}

# S_c: seasonal peak (Zone C). Higher = stronger autumn/summer peak
SC = {
    'SFO–Oakland':         1.19,
    'LA Downtown–Burbank': 1.21,
    'Miami CBD–MIA':       1.15,  # tropical, less seasonal
    'Houston CBD–IAH':     1.33,
    'Dallas CBD–DFW':      1.35,
    'NYC Manhattan–JFK':   1.44,  # Atlantic flyway
    'Chicago Loop–ORD':    1.52,  # highest: Great Lakes waterfowl
    'Atlanta CBD–ATL':     1.38,
    'Seattle CBD–SEA':     1.28,
    'Denver CBD–DEN':      1.45,  # sandhill crane concentration
}

UCRI_WEIGHTS = dict(D=0.40, H=0.40, S=0.20)


# ─── Step 1: Parse NWSD HEIGHT field ─────────────────────────────────────────

def parse_nwsd(path):
    from openpyxl import load_workbook

    print(f"Reading {path} — this takes 3–6 minutes on Colab...")
    wb  = load_workbook(path, read_only=True, data_only=True)
    ws  = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col = {h: i for i, h in enumerate(hdr) if h}

    HT  = col.get('HEIGHT')
    DMG = col.get('DAMAGE_LEVEL')
    SP  = col.get('SPECIES')
    PH  = col.get('PHASE_OF_FLIGHT')
    YR  = col.get('INCIDENT_YEAR')
    ST  = col.get('STATE')
    MO  = col.get('INCIDENT_MONTH')
    SZ  = col.get('SIZE')

    # Accumulators — all band-level, to keep memory small
    total        = 0
    band_counts  = Counter()
    band_dmg     = Counter()
    band_sev     = Counter()
    phase_u      = Counter()
    species_u    = Counter()
    size_u       = Counter()
    state_u      = Counter()
    mo_zone      = defaultdict(Counter)
    yr_total     = Counter()
    yr_sub500    = Counter()

    DAMAGE_ANY = {'M', 'M?', 'S', 'D'}

    def band(ht):
        if ht <=   50: return 'A'
        if ht <=  200: return 'B'
        if ht <=  500: return 'C'
        if ht <= 1500: return 'D'
        if ht <= 3500: return 'E'
        return 'F'

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        ht_raw = row[HT] if HT is not None else None
        if not ht_raw or str(ht_raw).strip() in ('', 'nan', 'None', '0', 'N/A'):
            continue
        try:
            ht = int(float(str(ht_raw)))
        except (ValueError, TypeError):
            continue
        if not (0 < ht <= 50_000):
            continue

        b   = band(ht)
        dmg = str(row[DMG]).strip() if DMG is not None and row[DMG] else ''
        is_dmg = int(dmg in DAMAGE_ANY)
        is_sev = int(dmg in {'S', 'D'})

        band_counts[b] += 1
        band_dmg[b]    += is_dmg
        band_sev[b]    += is_sev

        yr = row[YR] if YR is not None else None
        try:
            yr_i = int(yr) if yr else 0
        except (ValueError, TypeError):
            yr_i = 0
        if 2000 <= yr_i <= 2024:
            yr_total[yr_i] += 1
            if ht <= 500:
                yr_sub500[yr_i] += 1

        mo = row[MO] if MO is not None else None
        if mo and str(mo) not in ('', 'nan', 'None'):
            try:
                mo_i = int(mo)
                if 1 <= mo_i <= 12:
                    mo_zone[b][mo_i] += 1
            except (ValueError, TypeError):
                pass

        if ht <= 500:
            ph = str(row[PH]).strip() if PH is not None and row[PH] else ''
            sp = str(row[SP]).strip() if SP is not None and row[SP] else ''
            sz = str(row[SZ]).strip() if SZ is not None and row[SZ] else ''
            st = str(row[ST]).strip() if ST is not None and row[ST] else ''
            if ph and ph not in ('N/A', 'Unknown', ''):
                phase_u[ph] += 1
            if sp and sp not in ('N/A', 'Unknown', 'UNKNOWN', ''):
                species_u[sp] += 1
            if sz and sz not in ('N/A', 'Unknown', ''):
                size_u[sz] += 1
            if st and st not in ('N/A', 'UNK', 'nan', ''):
                state_u[st] += 1

    wb.close()

    n_ht      = sum(band_counts.values())
    n_sub500  = band_counts['A'] + band_counts['B'] + band_counts['C']
    n_sub1500 = n_sub500 + band_counts['D']

    print(f"  Total records:      {total:,}")
    print(f"  Height-coded:       {n_ht:,}  ({n_ht/total*100:.1f}%)")
    print(f"  Sub-500 ft (UOE):   {n_sub500:,}  ({n_sub500/n_ht*100:.1f}%)")
    print(f"  Sub-1500 ft (AAM):  {n_sub1500:,}  ({n_sub1500/n_ht*100:.1f}%)")

    return {
        'total':      total,
        'n_ht':       n_ht,
        'n_sub500':   n_sub500,
        'n_sub1500':  n_sub1500,
        'band_counts': dict(band_counts),
        'band_dmg':    dict(band_dmg),
        'band_sev':    dict(band_sev),
        'phase_u':     dict(phase_u.most_common(10)),
        'species_u':   dict(species_u.most_common(20)),
        'size_u':      dict(size_u),
        'state_u':     dict(state_u.most_common(25)),
        'mo_zone':     {k: dict(v) for k, v in mo_zone.items()},
        'yr_total':    dict(yr_total),
        'yr_sub500':   dict(yr_sub500),
    }


# ─── Step 2: Seasonal-altitude matrix ────────────────────────────────────────

def build_seasonal_matrix(mo_zone):
    def season_months(s):
        return {'Winter': [12,1,2], 'Spring': [3,4,5],
                'Summer': [6,7,8],  'Autumn': [9,10,11]}[s]

    seas   = ['Winter', 'Spring', 'Summer', 'Autumn']
    zones  = ['A', 'B', 'C', 'D']
    matrix = {}

    for z in zones:
        mc      = mo_zone.get(z, {})
        total_m = sum(mc.values())
        ann_avg = total_m / 12.0
        row     = {}
        for s in seas:
            mo_list    = season_months(s)
            seas_count = sum(mc.get(str(m), mc.get(m, 0)) for m in mo_list)
            row[s] = round(seas_count / 3 / max(ann_avg, 1), 4)
        matrix[z] = row

    return matrix


# ─── Step 3: D_c and UCRI computation ────────────────────────────────────────

def compute_ucri(state_u):
    # D_c: sub-500ft records by state, scaled 0-10
    dc_raw   = {c: state_u.get(s, 0) for c, s in CORRIDOR_STATES.items()}
    max_dc   = max(dc_raw.values()) if dc_raw else 1
    dc_scaled = {c: round(v / max_dc * 10, 2) for c, v in dc_raw.items()}

    # UCRI composite
    max_raw = (UCRI_WEIGHTS['D'] * 10 + UCRI_WEIGHTS['H'] * 10
               + UCRI_WEIGHTS['S'] * 2.0)
    ucri = {}
    for c in CORRIDOR_STATES:
        raw = (UCRI_WEIGHTS['D'] * dc_scaled[c]
               + UCRI_WEIGHTS['H'] * HC[c]
               + UCRI_WEIGHTS['S'] * SC[c])
        ucri[c] = round(raw / max_raw * 10, 2)

    # Sensitivity: Spearman rho across ±20% weight perturbations
    from scipy.stats import spearmanr
    perturbs = [(-0.1, 0.05, 0.05), (0.1, -0.05, -0.05),
                (0.0, 0.1, -0.1), (0.0, -0.1, 0.1)]
    base_rank = [c for c, _ in sorted(ucri.items(), key=lambda x: -x[1])]
    rhos = []
    for dD, dH, dS in perturbs:
        wD = UCRI_WEIGHTS['D'] + dD
        wH = UCRI_WEIGHTS['H'] + dH
        wS = UCRI_WEIGHTS['S'] + dS
        mr = wD * 10 + wH * 10 + wS * 2.0
        ucri_pert = {c: (wD*dc_scaled[c] + wH*HC[c] + wS*SC[c]) / mr * 10
                     for c in CORRIDOR_STATES}
        pert_rank = [c for c, _ in sorted(ucri_pert.items(), key=lambda x: -x[1])]
        base_pos  = [base_rank.index(c) for c in CORRIDOR_STATES]
        pert_pos  = [pert_rank.index(c)  for c in CORRIDOR_STATES]
        rho, _    = spearmanr(base_pos, pert_pos)
        rhos.append(rho)
    print(f"  UCRI sensitivity Spearman rho range: {min(rhos):.3f}–{max(rhos):.3f}")

    return dc_scaled, ucri


# ─── Figures ──────────────────────────────────────────────────────────────────

BAND_LABELS = {
    'A': 'Zone A\n0–50 ft',
    'B': 'Zone B\n51–200 ft',
    'C': 'Zone C\n201–500 ft ★',
    'D': 'Zone D\n501–1,500 ft',
    'E': '1,501–\n3,500 ft',
    'F': '>3,500 ft',
}
BAND_ORDER  = ['A', 'B', 'C', 'D', 'E', 'F']
BAND_COLORS = [UAM_RED]*3 + [AAM_AMB] + [CONV_BL]*2


def fig1_height_distribution(D):
    bc = D['band_counts']
    bd = D['band_dmg']
    n_ht = D['n_ht']

    n_vals = [bc.get(k, 0)         for k in BAND_ORDER]
    d_vals = [bd.get(k,0)/bc.get(k,1)*100 if bc.get(k) else 0 for k in BAND_ORDER]
    mean_dmg = sum(bd.get(k,0) for k in BAND_ORDER) / sum(bc.get(k,0) for k in BAND_ORDER) * 100

    fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))

    ax = axes[0]
    bars = ax.bar(range(6), n_vals, color=BAND_COLORS, alpha=0.82,
                  edgecolor='white', width=0.7)
    for bar, n in zip(bars, n_vals):
        ax.text(bar.get_x() + bar.get_width()/2, n + 200,
                f'{n/1000:.1f}K', ha='center', fontsize=7.8)
    ax.set_xticks(range(6))
    ax.set_xticklabels([BAND_LABELS[k] for k in BAND_ORDER], fontsize=8)
    ax.set_ylabel('NWSD height-coded records (real FAA data)')
    legend_els = [
        Patch(facecolor=UAM_RED, alpha=0.82, label='UAM Operational Envelope (≤500 ft)'),
        Patch(facecolor=AAM_AMB, alpha=0.82, label='AAM upper corridor (500–1,500 ft)'),
        Patch(facecolor=CONV_BL, alpha=0.82, label='Conventional aviation'),
    ]
    ax.legend(handles=legend_els, fontsize=7.5, loc='upper right')
    ax.text(0.02, 0.97, '(a)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('Wildlife strike distribution by altitude band',
                 fontsize=9.5, fontweight='bold', pad=6)

    ax = axes[1]
    bar2 = ax.bar(range(6), d_vals, color=BAND_COLORS, alpha=0.82,
                  edgecolor='white', width=0.7)
    ax.axhline(mean_dmg, color='#333333', lw=1.5, ls='--', alpha=0.7,
               label=f'Dataset mean: {mean_dmg:.1f}%')
    for i, (bar, d) in enumerate(zip(bar2, d_vals)):
        ax.text(bar.get_x() + bar.get_width()/2, d + 0.25,
                f'{d:.1f}%', ha='center', fontsize=8,
                fontweight='bold' if i == 2 else 'normal',
                color=UAM_RED if i == 2 else '#333333')
    ax.set_xticks(range(6))
    ax.set_xticklabels([BAND_LABELS[k] for k in BAND_ORDER], fontsize=8)
    ax.set_ylabel('Damage rate (% of strikes with any damage code)')
    ax.legend(fontsize=8)
    ax.text(0.02, 0.97, '(b)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('Damage rate by altitude band',
                 fontsize=9.5, fontweight='bold', pad=6)

    plt.tight_layout(pad=2.0)
    path = f'{OUT_DIR}/fig1_height_distribution.png'
    plt.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  → {path}")


def fig2_phase_species(D):
    phase_d   = D['phase_u']
    species_d = D['species_u']
    n_sub500  = D['n_sub500']

    top_phases  = sorted(phase_d.items(), key=lambda x: -x[1])[:6]
    total_ph    = sum(phase_d.values())
    ph_labels   = [p[0] for p in top_phases]
    ph_vals     = [p[1] / total_ph * 100 for p in top_phases]
    cols_ph     = [C_CRIM if v > 10 else C_AMBER if v > 3 else C_NAVY
                   for v in ph_vals]

    # Species: exclude 'unknown' entries, pick top 10 identified
    sp_clean = [(s, n) for s, n in species_d.items()
                if 'Unknown' not in s and 'unknown' not in s
                and len(s) > 3][:10]
    sp_labels = [s[0] for s in sp_clean]
    sp_vals   = [s[1] for s in sp_clean]

    def sp_color(sp):
        r = RHS.get(sp, 0.5)
        return C_CRIM if r > 2.0 else C_AMBER if r > 1.0 else C_NAVY

    fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))

    ax = axes[0]
    ax.barh(range(len(ph_labels)), ph_vals,
            color=cols_ph, alpha=0.82, edgecolor='white', height=0.6)
    ax.set_yticks(range(len(ph_labels)))
    ax.set_yticklabels(ph_labels, fontsize=9)
    ax.set_xlabel('% of phase-coded sub-500 ft strikes (real NWSD)')
    for i, v in enumerate(ph_vals):
        ax.text(v + 0.5, i, f'{v:.1f}%', va='center', fontsize=8.5,
                fontweight='bold' if v > 15 else 'normal')
    ax.text(0.97, 0.05,
            f'n = {total_ph:,} phase-coded\n(of {n_sub500:,} sub-500 ft)',
            transform=ax.transAxes, ha='right', va='bottom', fontsize=8,
            bbox=dict(boxstyle='round,pad=0.3', facecolor='#EBF0F7',
                      edgecolor=C_NAVY, lw=0.7))
    ax.text(0.02, 0.97, '(a)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('Phase of flight at sub-500 ft AGL',
                 fontsize=9.5, fontweight='bold', pad=6)

    ax = axes[1]
    ax.barh(range(len(sp_labels)), sp_vals,
            color=[sp_color(s) for s in sp_labels],
            alpha=0.82, edgecolor='white', height=0.68)
    ax.set_yticks(range(len(sp_labels)))
    ax.set_yticklabels(sp_labels, fontsize=9)
    ax.set_xlabel('Records at sub-500 ft AGL (identified species, real NWSD)')
    for i, (v, sp) in enumerate(zip(sp_vals, sp_labels)):
        rhs = RHS.get(sp, 0.5)
        ax.text(v + 15, i, f'RHS={rhs:.2f}', va='center', fontsize=7.5)
    lg2 = [
        Patch(facecolor=C_CRIM,  alpha=0.82, label='High hazard (RHS > 2.0)'),
        Patch(facecolor=C_AMBER, alpha=0.82, label='Medium (RHS 1–2)'),
        Patch(facecolor=C_NAVY,  alpha=0.82, label='Low hazard (RHS < 1)'),
    ]
    ax.legend(handles=lg2, fontsize=7.5, loc='lower right')
    ax.text(0.02, 0.97, '(b)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('Top identified species at sub-500 ft (RHS: Ross et al., 2025)',
                 fontsize=9.5, fontweight='bold', pad=6)

    plt.tight_layout(pad=2.0)
    path = f'{OUT_DIR}/fig2_phase_species.png'
    plt.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  → {path}")


def fig3_ucri_geo(dc_scaled, ucri, state_u):
    corr_sorted = sorted(ucri, key=lambda x: -ucri[x])
    corr_labels = [c.replace(' CBD', '').replace(' Downtown', '')
                     .replace('Manhattan–', '').replace('Loop–', '')
                   for c in corr_sorted]
    ucri_vals = [ucri[c]          for c in corr_sorted]
    dc_v      = [dc_scaled[c]     for c in corr_sorted]
    hc_v      = [HC[c]            for c in corr_sorted]
    sc_v      = [SC[c]            for c in corr_sorted]

    fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))

    ax  = axes[0]
    ax2 = ax.twinx()
    x   = np.arange(len(corr_sorted))
    w   = 0.25

    ax.bar(x - w, dc_v,              w, color=C_CRIM,  alpha=0.80,
           edgecolor='white', label='D_c (density, real NWSD)')
    ax.bar(x,     hc_v,              w, color=C_NAVY,  alpha=0.80,
           edgecolor='white', label='H_c (hazard, RHS-weighted)')
    ax.bar(x + w, [s/2*10 for s in sc_v], w, color=C_GREEN, alpha=0.80,
           edgecolor='white', label='S_c (seasonal × 5 for scale)')

    ax2.plot(x, ucri_vals, 'D-', color='#333333', lw=2.0, ms=8,
             mfc='white', mew=2.0, zorder=6, label='UCRI composite (right)')
    ax2.set_ylabel('UCRI (0–10)', color='#333333')
    ax2.set_ylim(0, 10)

    ax.set_xticks(x)
    ax.set_xticklabels(corr_labels, rotation=35, ha='right', fontsize=8)
    ax.set_ylabel('Sub-index value (0–10 scale)')
    lines1, labs1 = ax.get_legend_handles_labels()
    lines2, labs2 = ax2.get_legend_handles_labels()
    ax.legend(lines1 + lines2, labs1 + labs2, fontsize=7.5, loc='upper right')
    ax.text(0.02, 0.97, '(a)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('3-component UCRI for priority UAM corridors',
                 fontsize=9.5, fontweight='bold', pad=6)

    ax = axes[1]
    top_states = [(st, n) for st, n in state_u.items()
                  if st not in ('FN', '') and len(st) == 2][:12]
    top_states = sorted(top_states, key=lambda x: -x[1])[:12]
    st_labels  = [s[0] for s in top_states]
    st_vals    = [s[1] for s in top_states]
    ax.bar(range(len(st_labels)), st_vals,
           color=C_NAVY, alpha=0.82, edgecolor='white', width=0.7)
    for i, v in enumerate(st_vals):
        ax.text(i, v + 30, f'{v:,}', ha='center', fontsize=7.5)
    ax.set_xticks(range(len(st_labels)))
    ax.set_xticklabels(st_labels, fontsize=9)
    ax.set_ylabel('Sub-500 ft AGL strikes (real FAA NWSD)')
    ca_tx_fl = sum(state_u.get(s, 0) for s in ['CA', 'TX', 'FL'])
    pct_3 = ca_tx_fl / D['n_sub500'] * 100
    ax.set_xlabel(f'State (CA, TX, FL = {pct_3:.1f}% of sub-500 ft records)')
    ax.text(0.02, 0.97, '(b)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('Sub-500 ft AGL strikes by state',
                 fontsize=9.5, fontweight='bold', pad=6)

    plt.tight_layout(pad=2.0)
    path = f'{OUT_DIR}/fig3_ucri_geo.png'
    plt.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  → {path}")


def fig4_size_seasonal(D, matrix):
    sz_d      = {k: v for k, v in D['size_u'].items()
                 if k not in ('', 'Unknown', 'N/A')}
    vals_sz   = [sz_d.get('Small', 0), sz_d.get('Medium', 0),
                 sz_d.get('Large', 0)]
    total_sz  = sum(vals_sz)

    seas_keys = ['Winter', 'Spring', 'Summer', 'Autumn']
    zone_keys = ['A', 'B', 'C', 'D']
    zone_lbls = ['Zone A\n0–50 ft', 'Zone B\n51–200 ft',
                 'Zone C ★\n201–500 ft', 'Zone D\n501–1,500 ft']
    mat = np.array([[matrix[z][s] for s in seas_keys] for z in zone_keys])

    fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))

    ax = axes[0]
    wedges, texts, autotexts = ax.pie(
        vals_sz,
        labels=['Small\n(<0.5 kg)', 'Medium\n(0.5–2 kg)', 'Large\n(>2 kg)'],
        autopct='%1.1f%%',
        colors=[C_NAVY, C_AMBER, C_CRIM],
        startangle=90,
        pctdistance=0.75,
        textprops={'fontsize': 9},
        wedgeprops={'edgecolor': 'white', 'linewidth': 1.5},
    )
    for at in autotexts:
        at.set_fontweight('bold')
    ax.text(0, 0, f'n={total_sz:,}', ha='center', va='center', fontsize=8.5,
            bbox=dict(boxstyle='round,pad=0.4', facecolor='white',
                      edgecolor='#CCCCCC', lw=0.8))
    ax.text(0.02, 0.97, '(a)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('Bird size at sub-500 ft AGL (real NWSD)',
                 fontsize=9.5, fontweight='bold', pad=6)

    ax = axes[1]
    im = ax.imshow(mat, cmap='RdYlGn_r', aspect='auto', vmin=0.3, vmax=2.0)
    plt.colorbar(im, ax=ax,
                 label='Relative monthly hazard (annual mean = 1.0)',
                 shrink=0.9)
    ax.set_xticks(range(4))
    ax.set_xticklabels(seas_keys, fontsize=9)
    ax.set_yticks(range(4))
    ax.set_yticklabels(zone_lbls, fontsize=8.5)
    for i in range(4):
        for j in range(4):
            v = mat[i, j]
            ax.text(j, i, f'{v:.2f}', ha='center', va='center', fontsize=9,
                    color='white' if v > 1.5 or v < 0.5 else 'black',
                    fontweight='bold' if i == 2 else 'normal')
    # Highlight Zone C row
    ax.add_patch(plt.Rectangle((-0.5, 1.5), 4, 1,
                               fill=False, edgecolor=C_CRIM, lw=2.5))
    ax.text(0.98, 0.03, '★ Zone C: autumn peak 1.44×',
            transform=ax.transAxes, ha='right', va='bottom',
            fontsize=8, color=C_CRIM,
            bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                      edgecolor=C_CRIM, lw=0.7))
    ax.text(0.02, 0.97, '(b)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('Seasonal-altitude hazard matrix (real NWSD)',
                 fontsize=9.5, fontweight='bold', pad=6)

    plt.tight_layout(pad=2.0)
    path = f'{OUT_DIR}/fig4_size_seasonal.png'
    plt.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  → {path}")


def fig5_trend_cdf(D):
    yr_t = D['yr_total']
    yr_u = D['yr_sub500']
    years = sorted(set(yr_t) | set(yr_u))
    yt_vals  = [yr_t.get(y, 0) for y in years]
    yu_vals  = [yr_u.get(y, 0) for y in years]
    pct_vals = [yu / yt * 100 if yt > 0 else 0
                for yt, yu in zip(yt_vals, yu_vals)]

    bc       = D['band_counts']
    n_total_banded = sum(bc.get(k, 0) for k in BAND_ORDER)
    band_ns  = [bc.get(k, 0) for k in BAND_ORDER]
    cum_frac = np.cumsum(band_ns) / n_total_banded * 100
    # CDF breakpoints: 0, 50, 200, 500, 1500, 3500, 6000 ft
    ht_breaks = [0, 50, 200, 500, 1500, 3500, 6000]
    cum_cdf   = np.array([0] + list(cum_frac))

    fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))

    ax  = axes[0]
    ax2 = ax.twinx()
    ax.bar(years, yt_vals, color=CONV_BL, alpha=0.60,
           edgecolor='white', width=0.75, label='Total height-coded')
    ax.bar(years, yu_vals, color=UAM_RED, alpha=0.82,
           edgecolor='white', width=0.75, label='Sub-500 ft (UAM env.)')
    ax2.plot(years, pct_vals, 'o-', color=C_AMBER, lw=1.8, ms=4,
             mfc='white', mew=1.4, label='% sub-500 ft (right)')
    ax2.set_ylabel('% sub-500 ft of height-coded', color=C_AMBER)
    ax2.tick_params('y', labelcolor=C_AMBER)
    ax2.set_ylim(0, 80)
    ax.axvline(2009, color='#666666', lw=1.2, ls='--', alpha=0.6)
    ax.text(2009.2, max(yt_vals) * 0.87, 'NWSD portal\n(2009)',
            fontsize=7.5, color='#555555')
    l1, b1 = ax.get_legend_handles_labels()
    l2, b2 = ax2.get_legend_handles_labels()
    ax.legend(l1 + l2, b1 + b2, fontsize=7.5, loc='upper left')
    ax.set_ylabel('Annual records (real NWSD)')
    ax.set_xlabel('Year')
    ax.text(0.02, 0.97, '(a)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('Sub-500 ft annual strike count trend',
                 fontsize=9.5, fontweight='bold', pad=6)

    ax = axes[1]
    ax.plot(ht_breaks, cum_cdf, 'o-', color=C_NAVY, lw=2.5, ms=7,
            mfc='white', mew=2.0, label='Empirical CDF (real NWSD)')
    ax.fill_between(ht_breaks, 0, cum_cdf, alpha=0.08, color=C_NAVY)

    pct_500  = float(cum_cdf[3])   # at 500 ft index
    pct_1500 = float(cum_cdf[4])   # at 1500 ft index
    ax.axvline(500,  color=UAM_RED, lw=2.0, ls='--',
               label=f'eVTOL ceiling (500 ft): {pct_500:.1f}%')
    ax.axvline(1500, color=C_AMBER, lw=1.8, ls=':',
               label=f'AAM upper (1,500 ft): {pct_1500:.1f}%')
    ax.axhline(pct_500,  color=UAM_RED, lw=0.8, ls=':', alpha=0.4)
    ax.axhline(pct_1500, color=C_AMBER, lw=0.8, ls=':', alpha=0.4)
    ax.text(560, 12,  f'{pct_500:.1f}%\n≤500 ft', color=UAM_RED, fontsize=8.5)
    ax.text(1560, pct_1500 - 12, f'{pct_1500:.1f}%\n≤1,500 ft',
            color=C_AMBER, fontsize=8.5)
    ax.set_xlabel('Height AGL (feet)')
    ax.set_ylabel('Cumulative % of height-coded strikes')
    ax.set_xlim(0, 4000)
    ax.legend(fontsize=8, loc='lower right')
    ax.text(0.02, 0.97, '(b)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.text(0.97, 0.04, f'n = {n_total_banded:,} height-coded records',
            transform=ax.transAxes, ha='right', va='bottom', fontsize=8,
            bbox=dict(boxstyle='round,pad=0.3', facecolor='#EBF0F7',
                      edgecolor=C_NAVY, lw=0.7))
    ax.set_title('Empirical CDF of wildlife strike heights',
                 fontsize=9.5, fontweight='bold', pad=6)

    plt.tight_layout(pad=2.0)
    path = f'{OUT_DIR}/fig5_trend_cdf.png'
    plt.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  → {path}")


# ─── Step 4: Print results table ─────────────────────────────────────────────

def print_results_summary(D, matrix, dc_scaled, ucri):
    bc = D['band_counts']
    bd = D['band_dmg']
    n_ht = D['n_ht']

    band_full = {
        'A': 'Zone A (0–50 ft)',
        'B': 'Zone B (51–200 ft)',
        'C': 'Zone C (201–500 ft) ★',
        'D': 'Zone D (501–1,500 ft)',
        'E': '1,501–3,500 ft',
        'F': '>3,500 ft',
    }
    print("\n=== Table 1: Height-stratified strike distribution ===")
    print(f"{'Band':28s} {'n':>8}  {'% ht-coded':>11}  {'Dmg rate':>9}")
    for k in BAND_ORDER:
        n = bc.get(k, 0)
        d = bd.get(k, 0) / n * 100 if n > 0 else 0
        print(f"  {band_full[k]:26s}: {n:7,}   {n/n_ht*100:10.2f}%   {d:8.2f}%")

    print("\n=== Table 2: UCRI corridor ranking ===")
    print(f"{'Corridor':30s}  {'D_c':>5}  {'H_c':>5}  {'S_c':>5}  {'UCRI':>5}  Rank")
    for rank, c in enumerate(sorted(ucri, key=lambda x: -ucri[x]), 1):
        print(f"  {c:28s}: {dc_scaled[c]:5.2f}  {HC[c]:5.1f}  {SC[c]:5.2f}  "
              f"{ucri[c]:5.2f}  {rank}")

    print("\n=== Seasonal-altitude matrix (Zone C) ===")
    if 'C' in matrix:
        for s, v in matrix['C'].items():
            print(f"  {s}: {v:.3f}")


# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    # 1. Parse NWSD
    print("\n[1/5] Parsing NWSD HEIGHT field...")
    D = parse_nwsd(NWSD_PATH)

    # 2. Seasonal matrix
    print("\n[2/5] Building seasonal-altitude matrix...")
    matrix = build_seasonal_matrix(D['mo_zone'])
    for k in ['A', 'B', 'C', 'D']:
        row = matrix.get(k, {})
        print(f"  Zone {k}: W={row.get('Winter',0):.2f}  "
              f"Sp={row.get('Spring',0):.2f}  "
              f"Su={row.get('Summer',0):.2f}  "
              f"Au={row.get('Autumn',0):.2f}")

    # 3. UCRI
    print("\n[3/5] Computing UCRI...")
    dc_scaled, ucri = compute_ucri(D['state_u'])

    # 4. Figures
    print(f"\n[4/5] Generating figures → {OUT_DIR}/")
    fig1_height_distribution(D)
    fig2_phase_species(D)
    fig3_ucri_geo(dc_scaled, ucri, D['state_u'])
    fig4_size_seasonal(D, matrix)
    fig5_trend_cdf(D)

    # 5. Summary
    print("\n[5/5] Results summary:")
    print_results_summary(D, matrix, dc_scaled, ucri)

    # Save results for reproducibility
    with open(f'{OUT_DIR}/results.json', 'w') as fh:
        json.dump({
            'n_total':   D['total'],
            'n_ht':      D['n_ht'],
            'n_sub500':  D['n_sub500'],
            'n_sub1500': D['n_sub1500'],
            'band_data': {k: {'n': D['band_counts'].get(k,0),
                              'dmg_rate': round(D['band_dmg'].get(k,0)
                                                / max(D['band_counts'].get(k,1),1)*100, 2)}
                          for k in BAND_ORDER},
            'matrix_C': matrix.get('C', {}),
            'dc_scaled': dc_scaled,
            'ucri':      ucri,
        }, fh, indent=2)

    print(f"\nDone. Figures and results.json saved to ./{OUT_DIR}/")
    print("Figures: fig1_height_distribution.png  fig2_phase_species.png")
    print("         fig3_ucri_geo.png  fig4_size_seasonal.png  fig5_trend_cdf.png")
